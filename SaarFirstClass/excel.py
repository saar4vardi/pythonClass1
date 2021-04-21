from openpyxl import *
from random import *
workbook = Workbook()
sheet = workbook.active
sheet.title = "1st sheet" #נותן שם לטבלה
sheet["B1"] = "Test1"
sheet["B2"] = "Test2"
sheet["B3"] = "Test3"
sheet["A1"] = "Test1"
sheet["A2"] = "Test2"
sheet["A3"] = "Test3"
# print(sheet["A1"].value) #מראה את הערך שבתוך התא
for row in sheet.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2, values_only=True):
    print(row)
print()
for column in sheet.iter_cols(min_row=1, max_row=3, min_col=1, max_col=2, values_only=True):
        print(column)
# cells = sheet["A1:A100"] # הכנסת ערכים בין 1 ל 20 לתור A
# for cell in cells:
#     cell[0].value = randint(1, 20)
# print(workbook.sheetnames)
# print(sheet.cell(row=1, column=2).value) #הדפסת ערך של תא על פי מיקום התא בטבלה
# for i in range(1, 100): # הכנסת ערכים בין 1 ל 100 לתור A
#     for x in range(1, 100):
#         sheet.cell(row=i, column=x).value = randint(0, 100)

workbook.save(filename="PythonTest.xlsx") #שמירה
